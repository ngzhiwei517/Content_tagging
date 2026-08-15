import ast
import io
import re
import unittest
from contextlib import contextmanager
from pathlib import Path
from typing import Dict
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from ugc_tagger.final_update2_adapter import tag_candidates


APP_PATH = Path(__file__).resolve().parents[1] / "app.py"
APP_SOURCE = APP_PATH.read_text(encoding="utf-8")
APP_TREE = ast.parse(APP_SOURCE)


def load_function(name, namespace):
    node = next(
        item for item in APP_TREE.body
        if isinstance(item, (ast.FunctionDef, ast.AsyncFunctionDef)) and item.name == name
    )
    module = ast.Module(body=[node], type_ignores=[])
    ast.fix_missing_locations(module)
    exec(compile(module, str(APP_PATH), "exec"), namespace)
    return namespace[name]


class ExportInputOrderTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        namespace = {
            "pd": pd,
            "Dict": Dict,
            "io": io,
            "re": re,
            "MARKETS": ["PH", "MY", "ID", "KR", "SG", "VN", "TH"],
        }
        namespace["safe_str"] = load_function("safe_str", namespace)
        namespace["display_empty"] = load_function("display_empty", namespace)
        namespace["normalize_market"] = load_function("normalize_market", namespace)
        namespace["display_market"] = load_function("display_market", namespace)
        namespace["to_excel_bytes"] = load_function("to_excel_bytes", namespace)
        namespace["safe_sheet_name"] = load_function("safe_sheet_name", namespace)
        cls.grouped_excel_bytes = staticmethod(load_function("grouped_excel_bytes", namespace))

        order_namespace = {
            "pd": pd,
            "Dict": Dict,
            "final_update2_normalize_url": lambda value: str(value).strip().lower(),
        }
        cls.restore_input_order = staticmethod(
            load_function("restore_batch_input_order_v68_96", order_namespace)
        )

    def test_download_contract_does_not_sort_final_csv(self):
        download_block = APP_SOURCE.split('section_title("Downloads"', 1)[1]
        self.assertIn("same order as the uploaded or pasted posts", download_block)
        final_frame_line = next(
            line for line in download_block.splitlines()
            if line.strip().startswith("final_df = export_base_df")
        )
        self.assertNotIn("sort_values", final_frame_line)

    def test_grouped_workbook_keeps_input_order(self):
        rows = pd.DataFrame([
            {"Source": "Pasted links", "Market": "MY", "Track": "A", "Link": "third", "Views": 3},
            {"Source": "Pasted links", "Market": "KR", "Track": "B", "Link": "first", "Views": 100},
            {"Source": "Pasted links", "Market": "MY", "Track": "A", "Link": "second", "Views": 50},
            {"Source": "Pasted links", "Market": "Indonesia", "Track": "C", "Link": "fourth", "Views": 25},
        ])
        workbook = load_workbook(io.BytesIO(self.grouped_excel_bytes(rows)), read_only=True)
        sheet = workbook["All Posts"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        link_col = headers.index("Link") + 1
        links = [sheet.cell(row=row, column=link_col).value for row in range(2, 6)]
        self.assertEqual(links, ["third", "first", "second", "fourth"])
        self.assertIn("Market ID", workbook.sheetnames)
        id_sheet = workbook["Market ID"]
        id_headers = [cell.value for cell in next(id_sheet.iter_rows(min_row=1, max_row=1))]
        market_col = id_headers.index("Market") + 1
        self.assertEqual(id_sheet.cell(row=2, column=market_col).value, "ID")

    def test_final_results_restore_batch_order_after_backfill(self):
        batch = pd.DataFrame({"Link": ["first", "second", "third", "fourth"]})
        results = pd.DataFrame([
            {"Link": "third", "Views": 300},
            {"Link": "first", "Views": 200},
            {"Link": "fourth", "Views": 100},
        ])

        ordered = self.restore_input_order(results, batch)

        self.assertEqual(ordered["Link"].tolist(), ["first", "third", "fourth"])

    def test_unknown_result_links_stay_last_in_their_existing_order(self):
        batch = pd.DataFrame({"Link": ["first", "second"]})
        results = pd.DataFrame([
            {"Link": "unknown-a"},
            {"Link": "second"},
            {"Link": "unknown-b"},
        ])

        ordered = self.restore_input_order(results, batch)

        self.assertEqual(
            ordered["Link"].tolist(),
            ["second", "unknown-a", "unknown-b"],
        )

    def test_backend_grouping_restores_interleaved_input_order(self):
        class FakeBackend:
            @staticmethod
            @contextmanager
            def gemini_model_context(_model):
                yield _model

            @staticmethod
            def run_pipeline(records, *_args, **_kwargs):
                return pd.DataFrame([
                    {
                        "tiktok_url": record["webVideoUrl"],
                        "Creative Type": "Others",
                        "Content Details": "Test row",
                        "validation_status": "accepted",
                    }
                    for record in records
                ])

        candidates = pd.DataFrame([
            {"Source": "Pasted links", "Market": "MY", "Track": "A", "Link": "https://www.tiktok.com/@one/video/7001"},
            {"Source": "Pasted links", "Market": "KR", "Track": "B", "Link": "https://www.tiktok.com/@two/video/7002"},
            {"Source": "Pasted links", "Market": "MY", "Track": "A", "Link": "https://www.tiktok.com/@three/video/7003"},
        ])
        records = [
            {"id": "7001", "webVideoUrl": candidates.loc[0, "Link"]},
            {"id": "7002", "webVideoUrl": candidates.loc[1, "Link"]},
            {"id": "7003", "webVideoUrl": candidates.loc[2, "Link"]},
        ]
        with patch("ugc_tagger.final_update2_adapter.load_backend", return_value=FakeBackend()):
            tagged = tag_candidates(candidates, records, "key", "token")
        self.assertEqual(tagged["Link"].tolist(), candidates["Link"].tolist())
        self.assertTrue((tagged["Gemini Model"] == "gemini-3.1-flash-lite").all())

    def test_row_result_callback_preserves_completed_rows_before_failure(self):
        class InterruptedBackend:
            @staticmethod
            @contextmanager
            def gemini_model_context(_model):
                yield _model

            @staticmethod
            def run_pipeline(records, *_args, **kwargs):
                callback = kwargs["on_row_done"]
                for position, record in enumerate(records[:2]):
                    callback(
                        position + 1,
                        len(records),
                        {
                            "tiktok_url": record["webVideoUrl"],
                            "Creative Type": "Others",
                            "Content Details": "Completed before interruption",
                            "validation_status": "accepted",
                            "tier_used": "tier1_cover",
                        },
                        "tier1_cover",
                    )
                raise RuntimeError("simulated interruption")

        candidates = pd.DataFrame([
            {
                "Source": "Pasted links",
                "Market": "SG",
                "Track": "A",
                "Link": f"https://www.tiktok.com/@creator/video/80{index}",
            }
            for index in range(3)
        ])
        records = [
            {
                "id": f"80{index}",
                "webVideoUrl": candidates.loc[index, "Link"],
            }
            for index in range(3)
        ]
        completed = []
        with patch(
            "ugc_tagger.final_update2_adapter.load_backend",
            return_value=InterruptedBackend(),
        ):
            with self.assertRaisesRegex(RuntimeError, "simulated interruption"):
                tag_candidates(
                    candidates,
                    records,
                    "key",
                    "token",
                    on_result=lambda position, row, tier: completed.append(
                        (position, row, tier)
                    ),
                )
        self.assertEqual([position for position, _, _ in completed], [0, 1])
        self.assertEqual(
            [row["Link"] for _, row, _ in completed],
            candidates.iloc[:2]["Link"].tolist(),
        )
        self.assertTrue(all(row["Gemini Called"] for _, row, _ in completed))


if __name__ == "__main__":
    unittest.main()
