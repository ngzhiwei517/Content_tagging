import ast
import io
import os
import re
import unittest
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


APP_PATH = Path(__file__).resolve().parents[1] / "app.py"
APP_SOURCE = APP_PATH.read_text(encoding="utf-8")
APP_TREE = ast.parse(APP_SOURCE)


def load_function(name, namespace):
    node = next(
        item for item in APP_TREE.body
        if isinstance(item, (ast.FunctionDef, ast.AsyncFunctionDef)) and item.name == name
    )
    module = ast.Module(body=[node], type_ignores=[])
    ast.fix_missing_locations(module)
    exec(compile(module, str(APP_PATH), "exec"), namespace)
    return namespace[name]


class SensitiveSelectionContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.route_sensitive = staticmethod(load_function(
            "_route_sensitive_for_selection_v56",
            {"pd": pd, "Tuple": Tuple},
        ))
        cls.removed_mask = staticmethod(load_function(
            "_removed_mask_v56",
            {"pd": pd},
        ))

    def sample_rows(self):
        return pd.DataFrame([
            {
                "Link": "https://www.tiktok.com/@sensitive/video/7000000000000000001",
                "Tier Used": "SENSITIVE_HUMAN_REVIEW",
                "Raw Apify Status": "OK",
                "Review Action": "",
                "Needs Review": True,
                "Validation Status": "review",
                "QA Priority": "Review",
                "Review Note": "",
                "QA Reason": "",
            },
            {
                "Link": "https://www.tiktok.com/@normal/video/7000000000000000002",
                "Tier Used": "tier1",
                "Raw Apify Status": "OK",
                "Review Action": "KEEP",
                "Needs Review": False,
                "Validation Status": "passed",
                "QA Priority": "Passed",
                "Review Note": "",
                "QA Reason": "",
            },
        ])

    def test_top_posts_keeps_sensitive_in_human_review(self):
        routed, count = self.route_sensitive(self.sample_rows(), "Top posts")
        self.assertEqual(count, 1)
        self.assertEqual(routed.loc[0, "Review Action"], "")
        self.assertTrue(bool(routed.loc[0, "Needs Review"]))
        self.assertEqual(routed.loc[0, "Validation Status"], "review")
        self.assertEqual(routed.loc[0, "QA Priority"], "High")
        self.assertFalse(bool(routed.loc[0, "Manual Metrics Required"]))
        self.assertEqual(routed.loc[0, "Creative Type"], "")
        self.assertEqual(routed.loc[0, "Narrative"], "")
        self.assertEqual(routed.loc[0, "Content Details"], "")
        self.assertEqual(
            routed.loc[0, "Metrics Unavailable"],
            "Views, Likes, Comments, Shares, Saves",
        )
        self.assertIn("tag it manually", routed.loc[0, "Review Note"])
        self.assertEqual(routed.loc[1, "Review Action"], "KEEP")

    def test_tag_every_link_keeps_sensitive_in_human_review(self):
        routed, count = self.route_sensitive(self.sample_rows(), "Tag every link")
        self.assertEqual(count, 1)
        self.assertEqual(routed.loc[0, "Review Action"], "")
        self.assertTrue(bool(routed.loc[0, "Needs Review"]))
        self.assertEqual(routed.loc[0, "Validation Status"], "review")
        self.assertIn("retained for human review", routed.loc[0, "QA Reason"])

    def test_sensitive_review_is_not_removed_but_unavailable_still_is(self):
        routed, _ = self.route_sensitive(self.sample_rows(), "Tag every link")
        self.assertFalse(bool(self.removed_mask(routed).iloc[0]))
        unavailable = pd.DataFrame([{
            "Review Action": "REMOVE",
            "Validation Status": "removed",
        }])
        self.assertTrue(bool(self.removed_mask(unavailable).iloc[0]))

    def test_replacement_loop_remains_top_posts_only(self):
        self.assertIn(
            'replace_unavailable = replace_unavailable and st.session_state.get("selection_mode", "Top posts") == "Top posts"',
            APP_SOURCE,
        )

    def test_replacement_is_always_on_for_top_posts(self):
        self.assertNotIn('"Replace unavailable or sensitive posts automatically"', APP_SOURCE)
        assignment = "st.session_state.replace_unavailable_posts = True"
        assignment_index = APP_SOURCE.index(assignment)
        nearby_source = APP_SOURCE[max(0, assignment_index - 260):assignment_index]
        self.assertIn('if st.session_state.selection_mode == "Top posts":', nearby_source)

    def test_restricted_review_has_no_ai_fallback_and_optional_metrics(self):
        self.assertIn("<strong>Manual tagging required.</strong>", APP_SOURCE)
        self.assertIn("TikTok did not provide media or metadata for AI analysis.", APP_SOURCE)
        self.assertIn('if restricted_manual_review:', APP_SOURCE)
        self.assertIn('st.expander("Fill missing metrics (optional)"', APP_SOURCE)
        self.assertIn('placeholder="Not available"', APP_SOURCE)
        self.assertIn("missing_metrics = missing_metric_names(row.to_dict())", APP_SOURCE)
        self.assertIn("manual_metric_inputs", APP_SOURCE)
        self.assertIn('"Total Engagement"] = sum(', APP_SOURCE)
        self.assertNotIn("Please enter the available post metrics before saving.", APP_SOURCE)


class SummaryCopyContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        namespace = {
            "List": List,
            "Tuple": Tuple,
            "esc": lambda value: str(value),
            "safe_str": lambda value: "" if value is None else str(value).strip(),
        }
        cls.summary_kpi_row = staticmethod(load_function("summary_kpi_row", namespace))
        cls.focus_cards = staticmethod(load_function("focus_cards", namespace))
        cls.section_title = staticmethod(load_function("section_title", namespace))

    def test_section_heading_has_title_without_explanation(self):
        market = self.section_title("Market Summary", "#10b981")
        kol = self.section_title("KOL Size Performance", "#14b8a6")
        self.assertIn("Market Summary", market)
        self.assertIn("KOL Size Performance", kol)
        self.assertNotIn("<span>", market)
        self.assertNotIn("<span>", kol)

    def test_removed_summary_copy_is_absent(self):
        removed = [
            "Focus the summary",
            "Marketing view for tagged posts",
            "Grouped by primary creative type",
            "Performance grouped by primary creative type",
            "Compare uploaded files and pasted-link batches",
            "Performance by market",
            "Reach and engagement by creator size in the current filtered view",
            "Performance by market and track",
            "Highest-performing tagged posts in the current view",
            "Final files and internal QA report",
        ]
        for phrase in removed:
            with self.subTest(phrase=phrase):
                self.assertNotIn(phrase, APP_SOURCE)

    def test_summary_filters_remain_available(self):
        for widget_key in [
            "summary_source_v28",
            "summary_market_v28",
            "summary_track_v28",
            "summary_type_v55",
            "summary_metric_v28",
            "summary_sort_v28",
        ]:
            with self.subTest(widget_key=widget_key):
                self.assertIn(widget_key, APP_SOURCE)

    def test_blank_card_explanations_are_not_rendered(self):
        kpi = self.summary_kpi_row([("Views", "38.7M", "", "kpi-blue")])
        focus = self.focus_cards([("Main market", "MY", "", "focus-green")])
        self.assertNotIn("class='hint'", kpi)
        self.assertNotIn("class='sub'", focus)

    def test_only_useful_summary_explanations_remain(self):
        for removed in [
            "In current view",
            "Total reach",
            "Provided or blank",
            "Most common market in current view",
            "Use filters to focus on one campaign",
        ]:
            with self.subTest(removed=removed):
                self.assertNotIn(removed, APP_SOURCE)
        for kept in [
            "Average per post",
            "Mean of each post's engagement rate",
            'best_perf_sub = f"{short_num(highest_views_total)} views"',
        ]:
            with self.subTest(kept=kept):
                self.assertIn(kept, APP_SOURCE)


class GroupedExportContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        namespace = {
            "Dict": Dict,
            "io": io,
            "os": os,
            "pd": pd,
            "re": re,
            "safe_str": lambda value: "" if pd.isna(value) else str(value).strip(),
            "display_market": lambda value: str(value).strip() or "Other",
            "display_empty": lambda value, fallback: str(value).strip() if str(value).strip() else fallback,
        }
        namespace["to_excel_bytes"] = load_function("to_excel_bytes", namespace)
        namespace["safe_sheet_name"] = load_function("safe_sheet_name", namespace)
        cls.grouped_excel_bytes = staticmethod(load_function("grouped_excel_bytes", namespace))

    def test_grouped_export_does_not_create_duplicate_source_tabs(self):
        rows = pd.DataFrame([
            {"Source": "first_upload.csv", "Market": "MY", "Track": "Track A", "Link": "https://tiktok.com/1"},
            {"Source": "second_upload.xlsx", "Market": "MY", "Track": "Track B", "Link": "https://tiktok.com/2"},
        ])
        workbook = load_workbook(io.BytesIO(self.grouped_excel_bytes(rows)), read_only=True)
        self.assertEqual(workbook.sheetnames, ["All Posts", "Market MY", "Links Only"])
        self.assertNotIn("first_upload", workbook.sheetnames)
        self.assertNotIn("second_upload", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
